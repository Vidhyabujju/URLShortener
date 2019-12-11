#!/usr/bin/env python
# coding: utf-8

# In[3]:

try:   
    from flask import Flask, request, abort, render_template
    from flask_cors import CORS
    from flask_restful import Resource, Api
    from openpyxl import load_workbook
    import urllib.request, string, random, webbrowser, pandas, logging
    from urllib.parse import urlparse
    import cProfile, pstats, math, io
    from datetime import datetime

except ImportError:
    print ("No module found")
    
ps = cProfile.Profile()

app = Flask(__name__)
CORS(app)
app.config['PROPAGATE_EXCEPTIONS'] = True
api = Api(app)

@app.route('/')
def render_static():
    return render_template('index.html')

@app.route('/statistics')
def statistics():
    return render_template('profiler.html')

@app.route('/<string:key>')
def redirect(key):
    return render_template('redirect.html')

logging.basicConfig(filename='urlMinifier.log', level=logging.NOTSET, format='%(asctime)s %(levelname)s %(name)s %(threadName)s : %(message)s')

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, truncate_sheet=False):
    writer = pandas.ExcelWriter(filename, engine='openpyxl')

    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        writer.book = load_workbook(filename)

        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)

        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass

    if startrow is None:
        startrow = 0

    df.to_excel(writer, sheet_name, startrow=startrow,index=False, header=False)

    writer.save()
    
def shorten_url():
    try:
        chars=string.digits + string.ascii_letters
        shorten_url = ''.join(random.sample(chars, 6))
        minifiedURL = "http://127.0.0.1:5000/"+shorten_url
        return minifiedURL
    except ValueError:
        print('Mismatch with decalaration and input' + e)

##def writetoCSV(csvName):
##
##    
##    result = io.StringIO()
##    pstats.Stats(ps,stream=result).print_stats()
##    result=result.getvalue()
##    # chop the string into a csv-like buffer
##    result='ncalls'+result.split('ncalls')[-1]
##    result='\n'.join([','.join(line.rstrip().split(None,5)) for line in result.split('\n')])
##    # save it to disk
## 
##    with open(csvName, 'w+') as f:
##        #f=open(result.rsplit('.')[0]+'.csv','w')
##        f.write(result)
##        f.close()

class URL_Minifier(Resource):

    ps.enable()
    def post(self):
        app.logger.info('Minifying the URL')
        url1 = request.args.to_dict()
        url2=''
        for k, v in url1.items() :
            v=v.replace("&","%26")
            v=v.replace(",","%2C")
            url2=url2+k+'='+v+'&'
        url3=url2[:-1]
        url=url3[4:]
        url=url.replace(" ","+")
        o = urlparse(url)
        if (o.scheme == "") :
            url = "https://"+url
        try:
            df4 = pandas.read_excel('check.xlsx')
            if getattr(df4,'Long').isin([url]).any() :
                app.logger.info('The input URL was already minified by us')
                minifier = df4.loc[df4['Long'] == url, 'Short'].item()
                app.logger.info('Redirecting to the minified URL')
                #urllib.request.urlopen(minifier)
                return "Minification déjà effectuée. URL minifiée ==" + minifier
            else :
                minifier = shorten_url()
                rediretion = minifier[:-6]+"api/"+minifier[22:]
                app.logger.info('Minification complete')
                df1 = pandas.DataFrame({"Date" : [datetime.date(datetime.now())],"Long" : [url],"Short" : [minifier],"Redirection" : [rediretion]})
                app.logger.info('Updating the Dataset')
                append_df_to_excel('check.xlsx',df1)
                app.logger.info('Dataset update complete')
                app.logger.info('Redirecting to the minified URL')
                #urllib.request.urlopen(minifier)
                return "Veuillez trouver l'URL minifiée ==" + minifier
        except Exception as e:
            app.logger.error('Exception occured when minifying the URL')
            print(e)

    ps.disable()
##    writetoCSV('profiler-output.csv')

class URL_Redirection(Resource):

    ps.enable()
    def post(self,key):
        try:
            app.logger.info('Redirecting to the long URL')
            df4 = pandas.read_excel('check.xlsx')
            url = "http://127.0.0.1:5000/api/"+key
            if getattr(df4, 'Redirection').isin([url]).any() :
                val = df4.loc[df4['Redirection'] == url, 'Long'].item()
                #webbrowser.open_new_tab(val)
                app.logger.info('Redirection to the long URL complete')
                return (val)
            else:
                return abort(404)
        except Exception as e:
            app.logger.info('Exception occured when redirecting to the long URL')
            print (e)

    ps.disable()
##    writetoCSV('profiler-output.csv')

api.add_resource(URL_Minifier, '/api/minify')  
api.add_resource(URL_Redirection, '/api/<string:key>')

if __name__ == '__main__':
    app.run(debug=True)


# In[ ]:




